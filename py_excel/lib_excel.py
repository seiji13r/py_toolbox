from openpyxl import load_workbook

def cleanString(inValue):
    '''
    Return a string trimmed and with only printable characters.
    '''
    inStr = str(inValue).strip()
    if inStr == "":
        return ""

    tmpStrList = list()
    for character in inStr:
        if str(character).isprintable():
            tmpStrList.append(character)

    if len(tmpStrList) > 0:
        resString = "".join(tmpStrList)
        return resString

def excelTabToPyDict(excelFile=None, sheetName="Sheet1", headerRowNumber=1, headerRenameDict=None, reqColumns=None):
    resDict = dict()
    wb = load_workbook(excelFile, read_only=True, data_only=True)
    ws = wb[0]
    if sheetName in wb.sheetnames:
        ws = wb[sheetName]

    headerList = list()
    for rowNum, row in enumerate(ws.rows, start=1):
        # Skip the rows before the actual header name
        if rowNum < headerRowNumber:
            continue

        # Capture and clean the data of the row, all values will be strings.
        wsRow = list()
        for cell in row:
            cellVal = cell.value
            if cellVal is not None:
                cellVal = cleanString(cellVal)
            else:
                cellVal = ""
            wsRow.append(cellVal)

        # Capture the Header
        if rowNum == headerRowNumber:
            headerList = wsRow.copy()
            continue

        # Rename the Header to map actual key value pairs
        if headerRenameDict not in [None, "None", "", {}]:
            headerListTmp = list()
            for label in headerList:
                newLabel = label
                if label in headerRenameDict.keys():
                    newLabel = headerRenameDict[label]
                headerListTmp.append(newLabel)
                headerList = headerListTmp.copy()

        tmpRowDict = dict()
        for label, value in zip(headerList, wsRow):
            if reqColumns is not None:
                tmpRowDict[label] = value
            else:
                if label in reqColumns:
                    tmpRowDict[label] = value

        resDict[rowNum] = tmpRowDict
    return resDict